# search_logic.py

import re
import logging
from openpyxl import load_workbook
from openpyxl.styles.colors import Color  # Color 객체를 import 해야 합니다.
import os
from .config import RELATIVE_OFFSETS

# --- 로깅 설정 (사용자님 코드 그대로) ---
log_dir = 'logs'
if not os.path.exists(log_dir):
    os.makedirs(log_dir)
logging.basicConfig(filename=os.path.join(log_dir, 'search_errors.log'),
                    level=logging.ERROR,
                    format='%(asctime)s - %(levelname)s - %(message)s')


# --- [핵심] 사용자님의 정확한 get_status_from_color 함수 ---
def get_status_from_color(color_obj) -> str:
    """셀의 색상 객체를 분석하여 데이터 상태 텍스트("최신" 등)로 변환합니다."""
    # color_obj가 fill 객체일 수 있으므로, 실제 Color 객체는 fgColor에 있습니다.
    if not color_obj or not hasattr(color_obj, 'fgColor'):
        return "미지정"

    actual_color = color_obj.fgColor
    if not isinstance(actual_color, Color): return "미지정"

    if actual_color.type == 'theme':
        if actual_color.theme == 6: return "최신"
        if actual_color.theme == 3: return "1년 경과"
        if actual_color.theme in [0, 1]: return "1년 이상 경과"
    elif actual_color.type == 'rgb':
        hex_color = actual_color.rgb.upper() if actual_color.rgb else "00000000"
        if hex_color == "FFE2EFDA": return "최신"
        if hex_color == "FFDDEBF7": return "1년 경과"
        if hex_color in ["FFFFFFFF", "00000000", "FFFDEDEC"]: return "1년 이상 경과"
    return "미지정"


# --- 다른 헬퍼 함수들 (사용자님 코드 그대로) ---
def clean_text(text):
    if not isinstance(text, str):
        return text
    cleaned_text = re.sub(r'[\s\x00-\x1f\x7f]+', ' ', text)
    return cleaned_text.strip()


def parse_amount(text_value):
    if text_value is None: return 0
    text_value = str(text_value).strip()
    if not text_value: return 0
    try:
        return int(text_value.replace(",", ""))
    except (ValueError, TypeError):
        return 0


# --- [핵심 1] 요약 상태를 계산하는 새로운 함수 ---
def get_summary_status(statuses_dict):
    # 기준이 되는 세 항목의 상태를 리스트로 만듭니다.
    key_statuses = [
        statuses_dict.get('시평', '미지정'),
        statuses_dict.get('3년 실적', '미지정'),
        statuses_dict.get('5년 실적', '미지정')
    ]

    # 우선순위에 따라 전체 상태를 결정합니다.
    if '1년 이상 경과' in key_statuses:
        return '1년 이상 경과'
    if '1년 경과' in key_statuses:
        return '1년 경과'
    if all(s == '최신' for s in key_statuses):
        return '최신'

    return '미지정'  # 그 외의 모든 경우



# --- 최종 find_and_filter_companies 함수 ---
def find_and_filter_companies(file_path, filters):
    value_wb, style_wb = None, None  # 변수를 미리 선언
    try:
        # [핵심] 값용(value_wb)과 스타일용(style_wb)으로 파일을 두 번 엽니다.
        value_wb = load_workbook(filename=file_path, data_only=True)
        style_wb = load_workbook(filename=file_path, data_only=False)

        all_companies = []
        target_sheet_names = []
        region_filter = filters.get('region')
        if region_filter and region_filter != '전체':
            if region_filter in value_wb.sheetnames:
                target_sheet_names.append(region_filter)
        else:
            target_sheet_names = value_wb.sheetnames

        for sheet_name in target_sheet_names:
            value_sheet = value_wb[sheet_name]
            style_sheet = style_wb[sheet_name]
            max_row = value_sheet.max_row

            for r_idx in range(1, max_row + 1):
                # ... (이하 모든 데이터 처리 로직은 이전과 동일합니다) ...
                first_cell_value = value_sheet.cell(row=r_idx, column=1).value
                if isinstance(first_cell_value, str) and "회사명" in first_cell_value.strip():
                    for c_idx in range(2, value_sheet.max_column + 1):
                        try:
                            company_name = value_sheet.cell(row=r_idx, column=c_idx).value
                            if not isinstance(company_name, str) or not company_name.strip():
                                continue

                            company_data = {"검색된 회사": clean_text(company_name)}
                            company_data['대표지역'] = sheet_name.strip()

                            company_statuses = {}
                            for item, offset in RELATIVE_OFFSETS.items():
                                target_row = r_idx + offset
                                if target_row <= max_row:
                                    value = value_sheet.cell(row=target_row, column=c_idx).value
                                    style_cell = style_sheet.cell(row=target_row, column=c_idx)
                                    status = get_status_from_color(style_cell.fill)
                                    if item in ["부채비율", "유동비율"] and isinstance(value, (int, float)):
                                        processed_value = value * 100
                                    else:
                                        processed_value = clean_text(value) if isinstance(value, str) else value
                                    company_data[item] = processed_value if processed_value is not None else ""
                                    company_statuses[item] = status
                                else:
                                    company_data[item] = "N/A"
                                    company_statuses[item] = "N/A"

                            company_data["데이터상태"] = company_statuses
                            company_data["요약상태"] = get_summary_status(company_statuses)
                            all_companies.append(company_data)
                        except Exception as e:
                            logging.error(f"'{sheet_name}' 시트 데이터 처리 중 오류: {e}")
                            continue

        # --- 필터링 로직 (이전과 동일) ---
        filtered_results = all_companies
        if filters.get('name'):
            search_name = filters['name'].lower()
            filtered_results = [comp for comp in filtered_results if search_name in str(comp.get("검색된 회사", "")).lower()]
        if filters.get('manager'):
            search_manager = filters['manager'].lower()
            filtered_results = [comp for comp in filtered_results if search_manager in str(comp.get("비고", "")).lower()]
        for key, field_name in [('sipyung', '시평'), ('3y', '3년 실적'), ('5y', '5년 실적')]:
            min_val, max_val = filters.get(f'min_{key}'), filters.get(f'max_{key}')
            if min_val is not None:
                filtered_results = [comp for comp in filtered_results if
                                    (val := parse_amount(str(comp.get(field_name)))) is not None and val >= min_val]
            if max_val is not None:
                filtered_results = [comp for comp in filtered_results if
                                    (val := parse_amount(str(comp.get(field_name)))) is not None and val <= max_val]

        return filtered_results

    except Exception as e:
        logging.error(f"엑셀 파일 열기 실패: {file_path}, 오류: {e}")
        return []
    finally:
        # --- [핵심] 에러가 발생하든 안 하든, 작업이 끝나면 무조건 파일을 닫습니다. ---
        if value_wb:
            value_wb.close()
        if style_wb:
            style_wb.close()









# def find_and_filter_companies(file_path, filters):
#     all_companies = []
#
#     try:
#         workbook = load_workbook(filename=file_path, data_only=False)
#     except Exception as e:
#         logging.error(f"엑셀 파일 열기 실패: {file_path}, 오류: {e}")
#         return [{"오류": f"파일 열기 오류: {e}"}]
#
#     for sheet_name in workbook.sheetnames:
#         sheet = workbook[sheet_name]
#         max_row = sheet.max_row
#         max_col = sheet.max_column
#
#         for r_idx, row_cells in enumerate(sheet.iter_rows(max_row=max_row, max_col=max_col)):
#             excel_row_num = r_idx + 1
#
#             first_cell_value = row_cells[0].value
#             if isinstance(first_cell_value, str) and "회사명" in first_cell_value.strip():
#
#                 for c_idx, company_header_cell in enumerate(row_cells[1:], start=1):
#                     # ▼▼▼▼▼ [핵심 수정] 개별 회사 데이터를 읽는 부분을 try-except로 감쌉니다 ▼▼▼▼▼
#                     try:
#                         excel_col_num = company_header_cell.column
#                         company_name = company_header_cell.value
#
#                         if not isinstance(company_name, str) or not company_name.strip():
#                             continue
#
#                         cleaned_company_name = clean_text(company_name)
#                         company_data = {"검색된 회사": cleaned_company_name}
#
#                         clean_region_name = sheet_name.strip().replace('[', '').replace(']', '')
#                         company_data['대표지역'] = clean_region_name
#
#                         for item, offset in RELATIVE_OFFSETS.items():
#                             target_row = excel_row_num + offset
#                             if target_row <= max_row:
#                                 cell = sheet.cell(row=target_row, column=excel_col_num)
#                                 value = cell.value
#
#                                 # 데이터 처리 로직 (기존과 동일)
#                                 if item in ["부채비율", "유동비율"]:
#                                     if isinstance(value, (int, float)):
#                                         processed_value = value * 100
#                                     elif isinstance(value, str):
#                                         try:
#                                             processed_value = float(value.replace('%', '').strip())
#                                         except (ValueError, TypeError):
#                                             processed_value = clean_text(value)
#                                     else:
#                                         processed_value = value
#                                 elif item == "신용평가":
#                                     if isinstance(value, str):
#                                         cleaned_value = value.strip()
#                                         normalized_value = " ".join(cleaned_value.split())
#                                         processed_value = normalized_value.replace(' ', '\n', 1)
#                                     else:
#                                         processed_value = value
#                                 else:
#                                     processed_value = clean_text(value) if isinstance(value, str) else value
#
#                                 company_data[item] = processed_value if processed_value is not None else ""
#                             else:
#                                 company_data[item] = "N/A"
#
#                         company_statuses = {}
#                         for item, offset in RELATIVE_OFFSETS.items():
#                             target_row = excel_row_num + offset
#                             if target_row <= max_row:
#                                 cell = sheet.cell(row=target_row, column=excel_col_num)
#                                 company_statuses[item] = get_status_from_color(cell.fill.fgColor if cell.fill else None)
#                             else:
#                                 company_statuses[item] = "범위 초과"
#                         company_data["데이터상태"] = company_statuses
#
#                         all_companies.append(company_data)
#
#                     except Exception as e:
#                         # 오류 발생 시 로그 파일에 기록하고 다음 회사로 넘어갑니다.
#                         error_msg = (f"'{sheet_name}' 시트의 {excel_row_num}행, {company_header_cell.column}열 "
#                                      f"데이터 처리 중 오류 발생. 회사명: '{company_name}'. 오류: {e}")
#                         print(f"[경고] {error_msg}")
#                         logging.error(error_msg)
#                         continue  # ★★★ 이 부분이 중요합니다 ★★★
#                     # ▲▲▲▲▲ [핵심 수정] 여기까지 ▲▲▲▲▲
#
#     if not all_companies:
#         return [{"오류": "엑셀 파일에서 업체 정보를 찾을 수 없습니다."}]
#
#     # --- 필터링 로직 (기존과 동일) ---
#     filtered_results = all_companies
#     if filters.get('name'):
#         search_name = filters['name'].lower()
#         filtered_results = [comp for comp in filtered_results if search_name in str(comp.get("검색된 회사", "")).lower()]
#
#     if filters.get('manager'):
#         search_manager = filters['manager'].lower()
#         filtered_results = [comp for comp in filtered_results if search_manager in str(comp.get("비고", "")).lower()]
#
#     if filters.get('region') and filters['region'] != "전체":
#         search_region = filters['region'].strip().replace('[', '').replace(']', '')
#         filtered_results = [comp for comp in filtered_results if search_region == comp.get('대표지역')]
#
#     for key, field_name in [('sipyung', '시평'), ('perf_3y', '3년 실적'), ('perf_5y', '5년 실적')]:
#         min_val, max_val = filters.get(f'min_{key}'), filters.get(f'max_{key}')
#         if min_val is not None:
#             filtered_results = [comp for comp in filtered_results if
#                                 (val := parse_amount(str(comp.get(field_name)))) is not None and val >= min_val]
#         if max_val is not None:
#             filtered_results = [comp for comp in filtered_results if
#                                 (val := parse_amount(str(comp.get(field_name)))) is not None and val <= max_val]
#
#     if not filtered_results:
#         return [{"오류": "주어진 조건에 맞는 업체를 찾을 수 없습니다."}]
#
#     return filtered_results
