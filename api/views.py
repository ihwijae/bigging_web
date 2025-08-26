# api/views.py

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.conf import settings
import os
from . import search_logic
from drf_yasg.utils import swagger_auto_schema
from drf_yasg import openapi
from openpyxl import load_workbook
from django.core.files.storage import FileSystemStorage


class CompanySearchView(APIView):
    """
    다양한 조건으로 협력업체를 검색하는 API
    """

    @swagger_auto_schema(
        manual_parameters=[
            openapi.Parameter('name', openapi.IN_QUERY, description="업체명", type=openapi.TYPE_STRING),
            openapi.Parameter('region', openapi.IN_QUERY, description="지역 (엑셀 시트 이름)", type=openapi.TYPE_STRING),
            openapi.Parameter('manager', openapi.IN_QUERY, description="담당자 이름 ('비고' 컬럼)", type=openapi.TYPE_STRING),
            openapi.Parameter('min_sipyung', openapi.IN_QUERY, description="최소 시평액", type=openapi.TYPE_NUMBER),
            openapi.Parameter('max_sipyung', openapi.IN_QUERY, description="최대 시평액", type=openapi.TYPE_NUMBER),
            openapi.Parameter('min_3y', openapi.IN_QUERY, description="최소 3년 실적", type=openapi.TYPE_NUMBER),
            openapi.Parameter('max_3y', openapi.IN_QUERY, description="최대 3년 실적", type=openapi.TYPE_NUMBER),
            openapi.Parameter('min_5y', openapi.IN_QUERY, description="최소 5년 실적", type=openapi.TYPE_NUMBER),
            openapi.Parameter('max_5y', openapi.IN_QUERY, description="최대 5년 실적", type=openapi.TYPE_NUMBER),
        ]
    )
    def get(self, request, *args, **kwargs):
        # 1. 프론트에서 보낸 파일 타입을 받습니다. (기본값: 'eung')
        file_type = request.query_params.get('file_type', 'eung')

        # --- ▼▼▼ 이 부분을 수정합니다 ▼▼▼ ---
        # 2. 업로드된 파일이 저장되는 MEDIA_ROOT를 기준으로 파일 경로를 동적으로 생성합니다.
        excel_file_path = os.path.join(settings.MEDIA_ROOT, 'excel', f"{file_type}.xlsx")

        # --- ▲▲▲ 여기까지 수정 ---

        # 3. URL 쿼리 파라미터에서 모든 필터 값을 가져옵니다.
        # (이하 필터 로직은 이전과 동일)
        def get_int(param_name):
            val = request.query_params.get(param_name)
            try:
                return int(float(val)) if val else None
            except (ValueError, TypeError):
                return None

        filters = {
            'name': request.query_params.get('name'),
            'region': request.query_params.get('region', '전체'),
            'manager': request.query_params.get('manager'),
            'min_sipyung': get_int('min_sipyung'),
            'max_sipyung': get_int('max_sipyung'),
            'min_3y': get_int('min_3y'),
            'max_3y': get_int('max_3y'),
            'min_5y': get_int('min_5y'),
            'max_5y': get_int('max_5y'),
        }
        filters = {k: v for k, v in filters.items() if v is not None and v != ''}

        if not os.path.exists(excel_file_path):
            # 이제 파일이 없으면 검색 결과도 없고, 상태 표시도 '파일 없음'으로 일치하게 됩니다.
            return Response([], status=status.HTTP_200_OK)

        try:
            results = search_logic.find_and_filter_companies(excel_file_path, filters)

            for company in results:
                if '검색된 회사' in company:
                    company['업체명'] = company['검색된 회사']

            return Response(results, status=status.HTTP_200_OK)

        except Exception as e:
            print(f"!!! 필터링 에러 발생: {e}")
            return Response({"error": f"검색 중 오류가 발생했습니다: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class GetSheetNamesView(APIView):
    """
    엑셀 파일의 모든 시트 이름을 가져오는 API
    """

    def get(self, request, *args, **kwargs):
        # --- ▼▼▼ 이 부분을 수정합니다 ▼▼▼ ---
        # 1. 프론트에서 보낸 파일 타입을 받습니다. (기본값: 'eung')
        file_type = request.query_params.get('file_type', 'eung')

        # 2. 파일 타입에 맞는 엑셀 파일 경로를 동적으로 생성합니다.
        excel_file_path = os.path.join(settings.MEDIA_ROOT, 'excel', f"{file_type}.xlsx")
        # --- ▲▲▲ 여기까지 수정 ---

        if not os.path.exists(excel_file_path):
            # 파일이 없어도 에러 대신 빈 리스트를 보내 프론트가 처리하도록 함
            return Response([], status=status.HTTP_200_OK)

        try:
            workbook = load_workbook(filename=excel_file_path, read_only=True, keep_links=False)
            sheet_names = workbook.sheetnames
            return Response(sheet_names, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({"error": f"시트 이름을 읽는 중 오류 발생: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        finally:
            # --- [핵심] 작업이 끝나면 무조건 파일을 닫습니다. ---
            if workbook:
                workbook.close()


class ExcelFileUploadView(APIView):
    """
    엑셀 파일을 서버에 업로드하는 API.
    파일 타입(eung, tongsin, sobang)에 따라 정해진 이름으로 저장합니다.
    """

    def post(self, request, *args, **kwargs):
        file_obj = request.data.get('file')
        file_type = request.data.get('type')  # 'eung', 'tongsin', 'sobang'

        if not file_obj or not file_type:
            return Response({"error": "파일과 타입이 필요합니다."}, status=status.HTTP_400_BAD_REQUEST)

        # media/excel/ 폴더에 저장하도록 경로 설정
        upload_dir = os.path.join(settings.MEDIA_ROOT, 'excel')

        # 파일 타입에 따라 파일명 고정 (예: eung.xlsx)
        file_name = f"{file_type}.xlsx"

        fs = FileSystemStorage(location=upload_dir)

        # 만약 같은 이름의 파일이 이미 있다면 덮어쓰기
        if fs.exists(file_name):
            fs.delete(file_name)

        fs.save(file_name, file_obj)

        return Response({"message": f"'{file_name}' 파일이 성공적으로 업로드되었습니다."}, status=status.HTTP_201_CREATED)


class CheckFileStatusView(APIView):
    """
    서버에 각 종류의 엑셀 파일이 존재하는지 확인하는 API
    """

    def get(self, request, *args, **kwargs):
        upload_dir = os.path.join(settings.MEDIA_ROOT, 'excel')

        file_types = ['eung', 'tongsin', 'sobang']
        # --- 2. 변수 이름을 status에서 file_statuses로 변경하여 충돌을 피합니다. ---
        file_statuses = {}

        for file_type in file_types:
            file_path = os.path.join(upload_dir, f"{file_type}.xlsx")
            file_statuses[file_type] = os.path.exists(file_path)

        # --- 3. 이제 status.HTTP_200_OK가 올바르게 작동합니다. ---
        return Response(file_statuses, status=status.HTTP_200_OK)